from dotenv import load_dotenv
import logging
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

from livekit import agents
from livekit.agents import (
    AgentSession,
    Agent,
    RunContext,
    function_tool,
    cli,
    WorkerOptions,
)
from livekit.plugins import deepgram, openai, cartesia, silero
from demo_prompts import SYSTEM_PROMPT, GREETING_PROMPT

load_dotenv(".env.local")
logger = logging.getLogger("dr-immobilien-demo")
logger.setLevel(logging.INFO)

# Excel file path
EXCEL_FILE = "kunden_daten.xlsx"

def save_to_excel(data_type: str, data: dict):
    """Save customer data to Excel file"""
    try:
        # Check if file exists
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        
        # Get or create sheet
        if data_type not in wb.sheetnames:
            ws = wb.create_sheet(data_type)
            # Add headers based on data type
            if data_type == "Emails":
                ws.append(["Datum", "Uhrzeit", "Email", "Name", "Interesse"])
            elif data_type == "Kontakte":
                ws.append(["Datum", "Uhrzeit", "Name", "Telefon", "Anliegen"])
        else:
            ws = wb[data_type]
        
        # Add timestamp
        now = datetime.now()
        data["Datum"] = now.strftime("%Y-%m-%d")
        data["Uhrzeit"] = now.strftime("%H:%M:%S")
        
        # Append data
        if data_type == "Emails":
            ws.append([data["Datum"], data["Uhrzeit"], data.get("email", ""), 
                      data.get("name", ""), data.get("interesse", "")])
        elif data_type == "Kontakte":
            ws.append([data["Datum"], data["Uhrzeit"], data.get("name", ""), 
                      data.get("telefon", ""), data.get("anliegen", "")])
        
        # Save file
        wb.save(EXCEL_FILE)
        logger.info(f"Daten gespeichert in {EXCEL_FILE}")
        return True
    except Exception as e:
        logger.error(f"Fehler beim Speichern in Excel: {e}")
        return False


class ImmobilienAgent(Agent):
    def __init__(self) -> None:
        super().__init__(instructions=SYSTEM_PROMPT)
    
    async def on_enter(self):
        """Called when agent becomes active - start the conversation"""
        await self.session.generate_reply(instructions=GREETING_PROMPT)
    
    @function_tool
    async def get_firmeninfos(self, context: RunContext) -> str:
        """Zeige Firmeninformationen von DR Immobilien"""
        return """DR Immobilien & Partners GmbH
Milchstraße 9, 85049 Ingolstadt

Öffnungszeiten:
Montag-Donnerstag: 08:30-17:30 Uhr
Freitag: 08:30-12:00 Uhr
Samstag-Sonntag: Geschlossen

E-Mail: info@richarz-immobilien.de"""
    
    @function_tool
    async def kontakt_aufnehmen(
        self,
        context: RunContext,
        name: str,
        telefon: str,
        anliegen: str,
    ) -> str:
        """
        Nimm Kontaktdaten eines Interessenten auf.
        
        Args:
            name: Name des Kunden
            telefon: Telefonnummer
            anliegen: Was möchte der Kunde (z.B. "Wohnung mieten", "Haus kaufen")
        """
        logger.info(f"Kontakt: {name}, Tel: {telefon}, Anliegen: {anliegen}")
        
        # Save to Excel
        save_to_excel("Kontakte", {
            "name": name,
            "telefon": telefon,
            "anliegen": anliegen
        })
        
        return f"Vielen Dank {name}! Wir melden uns bei Ihnen unter {telefon} bezüglich '{anliegen}'."
    
    @function_tool
    async def email_speichern(
        self,
        context: RunContext,
        email: str,
        name: str = "",
        interesse: str = "",
    ) -> str:
        """
        Speichere die E-Mail-Adresse des Interessenten.
        
        Args:
            email: E-Mail-Adresse des Kunden
            name: Name des Kunden (optional)
            interesse: Welche Wohnung interessiert (optional, z.B. "3-Zimmer Zentrum")
        """
        logger.info(f"E-Mail gespeichert: {email}, Name: {name}, Interesse: {interesse}")
        
        # Save to Excel file
        save_to_excel("Emails", {
            "email": email,
            "name": name,
            "interesse": interesse
        })
        
        if interesse:
            return f"Perfekt! Ich sende Ihnen die Details zur {interesse} an {email}."
        else:
            return f"Perfekt! Ich sende Ihnen die Informationen an {email}."


async def entrypoint(ctx: agents.JobContext):
    """Main entrypoint for the demo voice agent"""
    logger.info(f"Connecting to room {ctx.room.name}")
    await ctx.connect()
    
    # Create agent session (Context7 Method 2: Plugin classes for language support)
    session = AgentSession(
        vad=silero.VAD.load(),
        stt=deepgram.STT(
            model="nova-3",
            language="de",  # German language
            detect_language=False,
        ),
        llm=openai.LLM(
            model="gpt-4o-mini",
            temperature=0.7,
        ),
        tts=cartesia.TTS(
            model="sonic-2",
            voice="b9de4a89-2257-424b-94c2-db18ba68c81a",  # German female voice
            language="de",  # German language for numbers and text
        ),
    )
    
    await session.start(
        room=ctx.room,
        agent=ImmobilienAgent(),
    )


if __name__ == "__main__":
    cli.run_app(WorkerOptions(entrypoint_fnc=entrypoint))
